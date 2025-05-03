from fastapi import FastAPI, UploadFile, File, HTTPException, Body, Request
from fastapi.responses import JSONResponse, Response, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import numpy as np
import json
import openpyxl
from typing import Dict, Any, List
import io
import os

# Import user and role routers
from app.api.users import router as users_router
from app.api.roles import router as roles_router

app = FastAPI(title="Excel to JSON Converter",
              description="API para convertir archivos Excel a JSON manteniendo espacios y celdas combinadas")

# Configurar CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Lista de orígenes permitidos (puedes usar ["*"] para permitir todos)
    allow_credentials=True,
    allow_methods=["*"],  # Permite todos los métodos
    allow_headers=["*"],  # Permite todos los headers
)

# Include routers
app.include_router(users_router, prefix="/api")
app.include_router(roles_router, prefix="/api")

# Initialize database with default admin role and user on startup
@app.on_event("startup")
async def startup_db_client():
    from app.db.connection import db
    from app.auth.security import get_password_hash
    from datetime import datetime
    
    # Create default admin role if it doesn't exist
    admin_role = await db.roles.find_one({"name": "admin"})
    if not admin_role:
        await db.roles.insert_one({
            "name": "admin",
            "permissions": ["read", "write", "delete", "admin"],
            "created_at": datetime.utcnow()
        })
    
    # Create default user role if it doesn't exist
    user_role = await db.roles.find_one({"name": "user"})
    if not user_role:
        await db.roles.insert_one({
            "name": "user",
            "permissions": ["read", "write"],
            "created_at": datetime.utcnow()
        })
    
    # Create default admin user if no users exist
    user_count = await db.users.count_documents({})
    if user_count == 0:
        await db.users.insert_one({
            "username": "admin",
            "email": "admin@example.com",
            "full_name": "Administrator",
            "hashed_password": get_password_hash("admin123"),
            "role": "admin",
            "created_at": datetime.utcnow()
        })
        print("Created default admin user: admin / admin123")

def process_excel(file_content: bytes, include_format: bool = False) -> Dict[str, Any]:
    """
    Procesa un archivo Excel y lo convierte a un diccionario JSON
    manteniendo la estructura de espacios y celdas combinadas.
    
    :param file_content: Contenido del archivo Excel en bytes
    :param include_format: Si es True, incluye información detallada de formato
    :return: Diccionario con los datos del Excel
    """
    try:
        if include_format:
            # Primero cargar con data_only=True para obtener valores calculados
            excel_file_with_values = io.BytesIO(file_content)
            workbook_with_values = openpyxl.load_workbook(excel_file_with_values, data_only=True)
            
            # Luego cargar sin data_only para obtener fórmulas
            excel_file_with_formulas = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file_with_formulas, data_only=False)
            
            result = {}
            
            # Procesar cada hoja
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                worksheet_with_values = workbook_with_values[sheet_name]
                sheet_data = []
                
                # Obtener información sobre celdas combinadas
                merged_cells = list(worksheet.merged_cells.ranges)
                merged_cell_map = {}
                
                for merged_range in merged_cells:
                    min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            # Marcar celdas que son parte de una combinación
                            merged_cell_map[(row, col)] = (min_row, min_col)
                
                # Determinar el rango de datos utilizados
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                for row_idx in range(1, max_row + 1):
                    row_data = []
                    
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell_with_value = worksheet_with_values.cell(row=row_idx, column=col_idx)
                        cell_key = (row_idx, col_idx)
                        
                        # Extraer valor de la celda (de la versión con valores calculados)
                        value = cell_with_value.value
                        
                        # Capturar la fórmula si existe (de la versión con fórmulas)
                        formula = None
                        if cell.data_type == 'f':  # 'f' indica que es una fórmula
                            formula = str(cell._value)  # La fórmula original
                        
                        # Obtener el formato de número para preservar porcentajes y moneda
                        number_format = cell.number_format
                        display_value = None
                        
                        # Detectar y mantener formatos especiales (porcentajes, moneda, etc.)
                        if isinstance(value, (int, float)):
                            # Para porcentajes
                            if "%" in number_format:
                                display_value = f"{value * 100}%"
                            # Para moneda
                            elif "$" in number_format or "€" in number_format or "¥" in number_format:
                                # Determinar el símbolo basado en el formato
                                symbol = "$"  # Default
                                if "€" in number_format:
                                    symbol = "€"
                                elif "¥" in number_format:
                                    symbol = "¥"
                                
                                # Formatear como moneda con el símbolo
                                display_value = f"{symbol}{value:,.2f}"
                        
                        # Si es parte de una celda combinada y no es la primera celda, usar None
                        if cell_key in merged_cell_map and merged_cell_map[cell_key] != cell_key:
                            value = None
                            display_value = None
                            formula = None
                        
                        # Extraer información de formato
                        format_info = {
                            "is_merged": cell_key in merged_cell_map,
                            "font": {
                                "name": cell.font.name,
                                "size": cell.font.size,
                                "bold": cell.font.bold,
                                "italic": cell.font.italic,
                                "color": cell.font.color.rgb if cell.font.color else None
                            },
                            "alignment": {
                                "horizontal": cell.alignment.horizontal,
                                "vertical": cell.alignment.vertical,
                                "wrap_text": cell.alignment.wrap_text
                            },
                            "fill": {
                                "color": cell.fill.start_color.rgb if cell.fill.start_color else None,
                                "pattern": cell.fill.fill_type
                            },
                            "border": {
                                "left": {"style": cell.border.left.style if cell.border.left else None},
                                "right": {"style": cell.border.right.style if cell.border.right else None},
                                "top": {"style": cell.border.top.style if cell.border.top else None},
                                "bottom": {"style": cell.border.bottom.style if cell.border.bottom else None}
                            },
                            "number_format": number_format
                        }
                        
                        # Si es una celda combinada, agregar información sobre el rango
                        if cell_key in merged_cell_map and merged_cell_map[cell_key] == cell_key:
                            for merged_range in merged_cells:
                                if merged_range.min_row == row_idx and merged_range.min_col == col_idx:
                                    format_info["merge_range"] = {
                                        "min_row": merged_range.min_row,
                                        "min_col": merged_range.min_col,
                                        "max_row": merged_range.max_row,
                                        "max_col": merged_range.max_col
                                    }
                                    break
                        
                        # Crear objeto de celda con valor, fórmula y formato
                        cell_obj = {
                            "value": value,
                            "display_value": display_value,
                            "format": format_info
                        }
                        
                        # Agregar la fórmula al objeto si existe
                        if formula:
                            cell_obj["formula"] = formula
                            # Asegurarse de que el valor calculado está presente
                            if value is None and cell_with_value.value is not None:
                                cell_obj["value"] = cell_with_value.value
                        
                        row_data.append(cell_obj)
                    
                    sheet_data.append(row_data)
                
                result[sheet_name] = sheet_data
            
            return result
        else:
            # Formato original simple usando pandas y openpyxl
            # Para las fórmulas y valores calculados, primero cargar con openpyxl
            excel_file = io.BytesIO(file_content)
            
            # Cargar dos versiones: una con valores calculados y otra con fórmulas
            workbook_with_values = openpyxl.load_workbook(excel_file, data_only=True)
            
            # Reiniciar el buffer para cargar de nuevo
            excel_file.seek(0)
            workbook_with_formulas = openpyxl.load_workbook(excel_file, data_only=False)
            
            # Recopilar formulas y valores calculados
            formulas = {}
            calculated_values = {}
            
            for sheet_name in workbook_with_formulas.sheetnames:
                sheet_formulas = workbook_with_formulas[sheet_name]
                sheet_values = workbook_with_values[sheet_name]
                
                sheet_formula_map = {}
                sheet_value_map = {}
                
                for row in sheet_formulas.iter_rows():
                    for cell in row:
                        coord = (cell.row, cell.column)
                        # Capturar fórmulas
                        if cell.data_type == 'f':
                            sheet_formula_map[coord] = str(cell._value)
                        
                        # Capturar el valor calculado correspondiente
                        value_cell = sheet_values.cell(row=cell.row, column=cell.column)
                        if value_cell.value is not None:
                            sheet_value_map[coord] = value_cell.value
                
                formulas[sheet_name] = sheet_formula_map
                calculated_values[sheet_name] = sheet_value_map
            
            # Reiniciar el buffer para leer con pandas
            excel_file.seek(0)
            excel_data = pd.read_excel(excel_file, sheet_name=None, header=None)
            
            result = {}
            
            # Procesar cada hoja
            for sheet_name, df in excel_data.items():
                # Reemplazar NaN con None para mantener espacios vacíos en el JSON
                df = df.replace({np.nan: None})
                
                # Convertir DataFrame a lista de listas para mantener estructura
                sheet_data = df.values.tolist()
                
                # Procesar filas para mantener tipos de datos correctos
                processed_data = []
                for row_idx, row in enumerate(sheet_data):
                    processed_row = []
                    for col_idx, cell in enumerate(row):
                        # Convertir tipos de datos Python a JSON compatibles
                        if isinstance(cell, (np.integer)):
                            cell_value = int(cell)
                        elif isinstance(cell, (np.floating)):
                            cell_value = float(cell)
                        elif isinstance(cell, (np.bool_)):
                            cell_value = bool(cell)
                        else:
                            cell_value = cell
                        
                        # Verificar si esta celda tiene una fórmula y/o valor calculado
                        coord = (row_idx + 1, col_idx + 1)  # +1 porque Excel es 1-indexed
                        has_formula = sheet_name in formulas and coord in formulas[sheet_name]
                        has_calc_value = sheet_name in calculated_values and coord in calculated_values[sheet_name]
                        
                        if has_formula:
                            # Crear un objeto con valor calculado y fórmula
                            cell_data = {
                                "formula": formulas[sheet_name][coord]
                            }
                            
                            # Priorizar el valor calculado de openpyxl si está disponible
                            if has_calc_value:
                                cell_value = calculated_values[sheet_name][coord]
                            
                            cell_data["value"] = cell_value
                            processed_row.append(cell_data)
                        else:
                            # Si no tiene fórmula, solo agregar el valor
                            processed_row.append(cell_value)
                    
                    processed_data.append(processed_row)
                
                result[sheet_name] = processed_data
            
            return result
    except Exception as e:
        import traceback
        trace = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo Excel: {str(e)}\n{trace}")

def json_to_excel(json_data: Dict[str, List[List[Any]]]) -> bytes:
    """
    Convierte datos JSON a un archivo Excel manteniendo la estructura y formato.
    
    :param json_data: Diccionario con nombres de hojas como claves y datos con formato
    :return: Archivo Excel en bytes
    """
    try:
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Eliminar la hoja por defecto que se crea
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Procesar cada hoja
        for sheet_name, sheet_data in json_data.items():
            # Crear una nueva hoja
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Procesar las celdas combinadas primero
            merged_ranges = []
            
            # Recorrer filas y columnas para aplicar valores y formato
            for row_idx, row in enumerate(sheet_data, 1):
                for col_idx, cell_data in enumerate(row, 1):
                    # Si es un diccionario con valor y formato
                    if isinstance(cell_data, dict) and "value" in cell_data:
                        cell_value = cell_data["value"]
                        display_value = cell_data.get("display_value")
                        cell_format = cell_data.get("format", {})
                        
                        # Si tenemos un valor para mostrar (porcentaje, moneda), usamos ese en su lugar
                        # pero guardamos el valor original para cálculos
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                        # Si hay una fórmula, establecerla en la celda
                        if "formula" in cell_data:
                            formula = cell_data["formula"]
                            if formula and formula.startswith("="):
                                try:
                                    # Establecer directamente la fórmula en la celda
                                    cell.value = None  # Limpiar el valor primero
                                    cell.data_type = "f"  # Indicar que es una fórmula
                                    cell._value = formula  # Asignar la fórmula bruta
                                    worksheet.formula_attributes[cell.coordinate] = {'t': 'shared', 'ref': cell.coordinate}
                                except Exception as e:
                                    # Si falla, mantener el valor calculado
                                    cell.value = cell_value
                                    # Añadir un comentario con la fórmula original
                                    comment = openpyxl.comments.Comment(f"Fórmula original: {formula}", "Sistema")
                                    cell.comment = comment
                        
                        # Aplicar formato de número si está disponible
                        if "number_format" in cell_format and cell_format["number_format"]:
                            cell.number_format = cell_format["number_format"]
                        
                        # Corregir específicamente para porcentajes y moneda
                        if display_value and isinstance(display_value, str):
                            # Para porcentajes
                            if display_value.endswith("%"):
                                try:
                                    percent_value = float(display_value.strip("%")) / 100
                                    cell.value = percent_value
                                    cell.number_format = "0%"
                                except ValueError:
                                    pass
                            # Para moneda
                            elif display_value.startswith(("$", "€", "¥")):
                                symbol = display_value[0]
                                try:
                                    value = float(display_value[1:].replace(",", ""))
                                    cell.value = value
                                    if symbol == "$":
                                        cell.number_format = '"$"#,##0.00'
                                    elif symbol == "€":
                                        cell.number_format = '"€"#,##0.00'
                                    elif symbol == "¥":
                                        cell.number_format = '"¥"#,##0.00'
                                except ValueError:
                                    pass
                        
                        # Aplicar formato de fuente
                        if "font" in cell_format:
                            font_data = cell_format["font"]
                            font_args = {}
                            
                            if font_data.get("name") is not None:
                                font_args["name"] = font_data.get("name")
                            if font_data.get("size") is not None:
                                font_args["size"] = font_data.get("size")
                            if font_data.get("bold") is not None:
                                font_args["bold"] = font_data.get("bold")
                            if font_data.get("italic") is not None:
                                font_args["italic"] = font_data.get("italic")
                            
                            # Manejar el color con cuidado
                            color_value = font_data.get("color")
                            if color_value is not None and isinstance(color_value, str):
                                font_args["color"] = openpyxl.styles.Color(rgb=color_value)
                            
                            cell.font = openpyxl.styles.Font(**font_args)
                        
                        # Aplicar alineación
                        if "alignment" in cell_format:
                            align_data = cell_format["alignment"]
                            align_args = {}
                            
                            if align_data.get("horizontal") is not None:
                                align_args["horizontal"] = align_data.get("horizontal")
                            if align_data.get("vertical") is not None:
                                align_args["vertical"] = align_data.get("vertical")
                            if align_data.get("wrap_text") is not None:
                                align_args["wrap_text"] = align_data.get("wrap_text")
                                
                            cell.alignment = openpyxl.styles.Alignment(**align_args)
                        
                        # Aplicar relleno (asegurarnos de manejar correctamente NULL y none)
                        if "fill" in cell_format:
                            fill_data = cell_format["fill"]
                            color_value = fill_data.get("color")
                            pattern = fill_data.get("pattern")
                            
                            # Solo aplicar si hay un color y es una cadena válida
                            if color_value is not None and isinstance(color_value, str) and color_value != "00000000":
                                fill_type = "solid"
                                if pattern and isinstance(pattern, str) and pattern != "none":
                                    fill_type = pattern
                                
                                fill = openpyxl.styles.PatternFill(
                                    fill_type=fill_type,
                                    start_color=color_value
                                )
                                cell.fill = fill
                        
                        # Aplicar bordes
                        if "border" in cell_format:
                            border_data = cell_format["border"]
                            border_sides = {}
                            
                            # Procesar cada lado del borde con cuidado
                            for side in ["left", "right", "top", "bottom"]:
                                if side in border_data and border_data[side] is not None:
                                    side_style = border_data[side].get("style")
                                    if side_style is not None and side_style != "none":
                                        border_sides[side] = openpyxl.styles.Side(style=side_style)
                            
                            # Crear el borde solo con los lados definidos
                            if border_sides:
                                cell.border = openpyxl.styles.Border(**border_sides)
                        
                        # Registrar celdas combinadas
                        if "merge_range" in cell_format:
                            merge_data = cell_format["merge_range"]
                            min_row = merge_data.get("min_row", row_idx)
                            min_col = merge_data.get("min_col", col_idx)
                            max_row = merge_data.get("max_row", row_idx)
                            max_col = merge_data.get("max_col", col_idx)
                            
                            cell_range = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:" \
                                        f"{openpyxl.utils.get_column_letter(max_col)}{max_row}"
                            
                            merged_ranges.append(cell_range)
                    else:
                        # Si es un valor simple (para compatibilidad con el formato anterior)
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_data)
            
            # Aplicar combinaciones de celdas
            for cell_range in merged_ranges:
                worksheet.merge_cells(cell_range)
            
            # Ajustar ancho de columnas
            for col_idx in range(1, worksheet.max_column + 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(col_idx)
                
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        import traceback
        trace = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"Error convirtiendo JSON a Excel: {str(e)}\n{trace}")

def generate_form_html(excel_data: Dict[str, Any]) -> str:
    """
    Genera el código HTML de un formulario basado en los datos de Excel.
    
    :param excel_data: Datos del Excel con información de formato
    :return: Código HTML del formulario
    """
    html = """
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Formulario basado en Excel</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
            }
            .excel-form {
                border-collapse: collapse;
                width: 100%;
            }
            .excel-form td {
                padding: 8px;
                border: 1px solid #ddd;
            }
            .excel-form input, .excel-form select {
                width: 100%;
                padding: 5px;
                box-sizing: border-box;
            }
            .bold-text {
                font-weight: bold;
            }
            .italic-text {
                font-style: italic;
            }
            .right-aligned {
                text-align: right;
            }
            .centered {
                text-align: center;
            }
            .submit-btn {
                margin-top: 20px;
                padding: 10px 15px;
                background-color: #4CAF50;
                color: white;
                border: none;
                cursor: pointer;
            }
            .tab-container {
                margin-top: 20px;
            }
            .tab {
                overflow: hidden;
                border: 1px solid #ccc;
                background-color: #f1f1f1;
            }
            .tab button {
                background-color: inherit;
                float: left;
                border: none;
                outline: none;
                cursor: pointer;
                padding: 10px 16px;
                transition: 0.3s;
            }
            .tab button:hover {
                background-color: #ddd;
            }
            .tab button.active {
                background-color: #ccc;
            }
            .tabcontent {
                display: none;
                padding: 6px 12px;
                border: 1px solid #ccc;
                border-top: none;
            }
        </style>
    </head>
    <body>
        <h1>Formulario basado en Excel</h1>
        <div class="tab-container">
            <div class="tab">
    """
    
    # Crear pestañas para cada hoja
    for i, sheet_name in enumerate(excel_data.keys()):
        active = " class='active'" if i == 0 else ""
        html += f"""<button id="tab-{i}" onclick="openSheet(event, 'sheet-{i}')"{active}>{sheet_name}</button>\n"""
    
    html += """
            </div>
    """
    
    # Crear el contenido de cada hoja
    for i, (sheet_name, sheet_data) in enumerate(excel_data.items()):
        display = "block" if i == 0 else "none"
        html += f"""
            <div id="sheet-{i}" class="tabcontent" style="display: {display};">
                <form id="form-{i}" class="excel-form-container">
                    <table class="excel-form">
        """
        
        # Generar filas y celdas del formulario
        for row_idx, row in enumerate(sheet_data):
            html += "<tr>\n"
            
            for col_idx, cell in enumerate(row):
                # Si la celda tiene formato
                if isinstance(cell, dict) and "value" in cell and "format" in cell:
                    value = cell["value"] if cell["value"] is not None else ""
                    display_value = cell.get("display_value", value)
                    cell_format = cell["format"]
                    
                    # Determinar si es celda combinada
                    is_merged = cell_format.get("is_merged", False)
                    colspan = ""
                    rowspan = ""
                    
                    # Si es la celda principal de una combinación
                    if "merge_range" in cell_format:
                        merge_range = cell_format["merge_range"]
                        colspan = f" colspan=\"{merge_range['max_col'] - merge_range['min_col'] + 1}\""
                        rowspan = f" rowspan=\"{merge_range['max_row'] - merge_range['min_row'] + 1}\""
                    
                    # Si no es una celda combinada secundaria
                    if not is_merged or (is_merged and "merge_range" in cell_format):
                        # Generar clases de estilo
                        classes = []
                        styles = []
                        
                        # Aplicar formato de fuente
                        if "font" in cell_format:
                            font = cell_format["font"]
                            if font.get("bold"):
                                classes.append("bold-text")
                            if font.get("italic"):
                                classes.append("italic-text")
                            if font.get("name"):
                                styles.append(f"font-family: {font['name']}")
                            if font.get("size"):
                                styles.append(f"font-size: {font['size']}pt")
                            if font.get("color"):
                                color = font["color"]
                                if isinstance(color, str) and len(color) >= 6:
                                    styles.append(f"color: #{color[2:]}")
                        
                        # Aplicar alineación
                        if "alignment" in cell_format:
                            align = cell_format["alignment"]
                            if align.get("horizontal") == "right":
                                classes.append("right-aligned")
                            elif align.get("horizontal") == "center":
                                classes.append("centered")
                        
                        # Aplicar color de fondo
                        if "fill" in cell_format and cell_format["fill"].get("color"):
                            color = cell_format["fill"]["color"]
                            if isinstance(color, str) and len(color) >= 6 and color != "00000000":
                                styles.append(f"background-color: #{color[2:]}")
                        
                        # Unir clases y estilos
                        class_attr = ""
                        if classes:
                            class_attr = f" class=\"{' '.join(classes)}\""
                        
                        style_attr = ""
                        if styles:
                            style_attr = f" style=\"{'; '.join(styles)}\""
                        
                        # Determinar si es encabezado o dato
                        is_header = row_idx == 0 or col_idx == 0 or (cell_format.get("font", {}).get("bold", False))
                        
                        # Generar HTML de la celda
                        html += f"<td{colspan}{rowspan}{class_attr}{style_attr}>"
                        
                        # Determinar el tipo de elemento del formulario
                        if is_header:
                            # Es encabezado, mostramos solo texto
                            html += f"{display_value if display_value is not None else value}"
                        else:
                            # Detectar tipo de dato para el input adecuado
                            input_type = "text"
                            input_value = display_value if display_value is not None else value
                            
                            # Para porcentajes
                            if isinstance(display_value, str) and display_value.endswith("%"):
                                input_type = "number"
                                input_value = display_value.rstrip("%")
                                
                            # Para moneda
                            elif isinstance(display_value, str) and display_value and display_value[0] in ("$", "€", "¥"):
                                input_type = "number"
                                input_value = display_value[1:].replace(",", "")
                                
                            # Para fechas
                            elif "date" in cell_format.get("number_format", "").lower():
                                input_type = "date"
                            
                            # ID único para el campo
                            field_id = f"field-{sheet_name}-{row_idx}-{col_idx}"
                            
                            html += f"<input type=\"{input_type}\" id=\"{field_id}\" name=\"{field_id}\" value=\"{input_value}\">"
                        
                        html += "</td>\n"
                else:
                    # Celda simple sin formato
                    value = cell if cell is not None else ""
                    html += f"<td>{value}</td>\n"
            
            html += "</tr>\n"
        
        html += """
                    </table>
                    <button type="submit" class="submit-btn">Enviar</button>
                </form>
            </div>
        """
    
    # Agregar JavaScript para la funcionalidad de pestañas y formulario
    html += """
        <script>
            function openSheet(evt, sheetId) {
                var i, tabcontent, tablinks;
                tabcontent = document.getElementsByClassName("tabcontent");
                for (i = 0; i < tabcontent.length; i++) {
                    tabcontent[i].style.display = "none";
                }
                tablinks = document.getElementsByClassName("tab")[0].getElementsByTagName("button");
                for (i = 0; i < tablinks.length; i++) {
                    tablinks[i].className = tablinks[i].className.replace(" active", "");
                }
                document.getElementById(sheetId).style.display = "block";
                evt.currentTarget.className += " active";
            }
            
            // Manejar envío de formulario
            document.addEventListener('DOMContentLoaded', function() {
                const forms = document.querySelectorAll('.excel-form-container');
                forms.forEach(form => {
                    form.addEventListener('submit', function(e) {
                        e.preventDefault();
                        
                        // Recopilar datos del formulario
                        const formData = new FormData(this);
                        const formValues = {};
                        
                        for (let [key, value] of formData.entries()) {
                            formValues[key] = value;
                        }
                        
                        // Aquí puedes enviar los datos mediante fetch a tu API
                        console.log('Datos del formulario:', formValues);
                        alert('Formulario enviado correctamente');
                    });
                });
            });
        </script>
    </body>
    </html>
    """
    
    return html

@app.post("/convert", response_class=JSONResponse)
async def convert_excel_to_json(file: UploadFile = File(...), include_format: bool = False):
    """
    Convierte un archivo Excel a formato JSON manteniendo la estructura,
    espacios y celdas combinadas.
    
    :param file: Archivo Excel a convertir
    :param include_format: Si es True, incluye información detallada de formato como tipo de letra, 
                          tamaño, estilo, colores, etc. Por defecto es False (formato simple).
    """
    # Verificar que el archivo sea Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser de tipo Excel (.xlsx o .xls)")
    
    # Leer el contenido del archivo
    contents = await file.read()
    
    # Procesar el Excel
    result = process_excel(contents, include_format)
    
    return result

@app.post("/convert-to-excel", response_class=Response)
async def convert_json_to_excel(data: Dict[str, List[List[Any]]] = Body(...)):
    """
    Convierte datos JSON a un archivo Excel manteniendo la estructura y formato.
    
    El formato esperado es un diccionario donde:
    - Las claves son nombres de hojas
    - Los valores son listas de listas, que pueden contener:
      - Valores simples (número, texto, etc.) 
      - O diccionarios con "value" y "format" para mantener el estilo
    
    Ejemplo simple:
    {
        "Hoja1": [
            ["Nombre", "Edad", "Profesión"],
            ["Juan", 30, "Ingeniero"]
        ]
    }
    
    Ejemplo con formato:
    {
        "Hoja1": [
            [
                {"value": "Nombre", "format": {"font": {"bold": true}}},
                {"value": "Edad", "format": {"font": {"bold": true}}}
            ],
            [
                {"value": "Juan", "format": {"font": {"color": "FF0000"}}},
                {"value": 30, "format": {"alignment": {"horizontal": "right"}}}
            ]
        ]
    }
    """
    try:
        # Validar estructura del JSON
        if not isinstance(data, dict):
            raise HTTPException(status_code=400, detail="Los datos deben ser un objeto JSON")
        
        for sheet_name, sheet_data in data.items():
            if not isinstance(sheet_data, list):
                raise HTTPException(status_code=400, 
                                  detail=f"Los datos para la hoja '{sheet_name}' deben ser una lista de filas")
            
            for row in sheet_data:
                if not isinstance(row, list):
                    raise HTTPException(status_code=400, 
                                      detail=f"Cada fila en la hoja '{sheet_name}' debe ser una lista de celdas")
        
        # Convertir JSON a Excel
        excel_bytes = json_to_excel(data)
        
        # Devolver el archivo Excel
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=converted_data.xlsx"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en la conversión: {str(e)}")

@app.post("/convert-to-form", response_class=HTMLResponse)
async def convert_excel_to_form(file: UploadFile = File(...)):
    """
    Convierte un archivo Excel a un formulario web HTML con estilos CSS
    que replican la apariencia del Excel original.
    """
    # Verificar que el archivo sea Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser de tipo Excel (.xlsx o .xls)")
    
    # Leer el contenido del archivo
    contents = await file.read()
    
    # Procesar el Excel con formato detallado
    excel_data = process_excel(contents, include_format=True)
    
    # Generar el HTML del formulario
    form_html = generate_form_html(excel_data)
    
    # Retornar el HTML completo
    return form_html

@app.get("/")
async def read_root():
    return {"message": "API para convertir entre Excel y JSON conservando formato. Usa los endpoints /convert y /convert-to-excel."}